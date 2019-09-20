# xlsx spreadsheet parser
# how about list of namedtuples as output parsing spreadsheet? pros: immutable, access by name, default values

import logging
import os
import sys

from openpyxl import load_workbook, Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill


from errors import BabelError
from data.data_objs import VenData
from data.validators import (shorten4datastore, value2string, normalize_date,
                             normalize_isbn, normalize_price,
                             normalize_whitespaces)
from logging_settings import format_traceback


mlogger = logging.getLogger('babel')


FONT_BOLD = Font(bold=True)
FILL_GRAY = PatternFill(fill_type='solid',
                        start_color='C4C5C6',
                        end_color='C4C5C6')


class SheetReader:

    def __init__(self, file):
        self.wb = load_workbook(
            filename=file,
            data_only=True)
        self.ws = self.wb.active
        self.max_row = 0
        self.max_column = 1
        for row in self.ws.rows:
            self.max_row += 1
            stripped_row = [x for x in row if x.value is not None]
            max_column = len(stripped_row)
            if max_column > self.max_column:
                self.max_column = max_column
        self.range = 'A1:' + str(
            get_column_letter(self.max_column)) + str(self.max_row)

        mlogger.debug(
            f'Loaded sheet {file}: data range detected: {self.range}')

    def __iter__(self):
        for row in self.ws[self.range]:
            data = []
            for cell in row:
                data.append(cell.value)
            yield data


class ResourceDataReader:
    """
    Parses specified rows and columns of the xlsx spreadsheet
    The reader can be iterater over in a loop

    Arguments:
    ----------
    fh: str
        path to xlsx spreadsheet
    header_row: int
        row number of a header, index starts with 0
    title_col: int
        title column number, index starts with 0
    add_title_col: int
        additional title column, index starts with 0
    author_col: int
        author column number, index starts with 0
    series_col: int
        series column number, index starts with 0
    publisher_col: int
        publisher column number, index starts with 0
    pub_datae_col: int
        publicattion column number, index starts with 0
    summary_col: int
        summary column number, index starts with 0
    isbn_col: int
        ISBN column number, index starts with 0
    upc_col: int
        UPC column number, index starts with 0
    other_no_col: int
        publisher interal ID, etc. column number, index starts with 0
    price_list_col: int
        list price column number, index starts with 0
    price_disc: int
        discounted price column number, index starts with 0
    desc_url: int
        URL of description column number, index starts with 0

    """

    def __init__(
            self, fh,
            header_row=None,
            title_col=None,
            add_title_col=None,
            author_col=None,
            series_col=None,
            publisher_col=None,
            pub_date_col=None,
            pub_place_col=None,
            summary_col=None,
            isbn_col=None,
            upc_col=None,
            other_no_col=None,
            price_list_col=None,
            price_disc_col=None,
            desc_url_col=None,
            misc_col=None):

        self.header_row = header_row
        self.title_col = title_col
        self.add_title_col = add_title_col
        self.author_col = author_col
        self.series_col = series_col
        self.publisher_col = publisher_col
        self.pub_date_col = pub_date_col
        self.pub_place_col = pub_place_col
        self.summary_col = summary_col
        self.isbn_col = isbn_col
        self.upc_col = upc_col
        self.other_no_col = other_no_col
        self.price_list_col = price_list_col
        self.price_disc_col = price_disc_col
        self.desc_url_col = desc_url_col
        self.misc_col = misc_col

        try:
            self.min_row = header_row + 1
        except TypeError:
            raise AttributeError(
                'Header row number is a required argument')

        if title_col is None:
            raise AttributeError(
                'Title column number is a required argument')

        wb = load_workbook(
            filename=fh,
            read_only=True,
            keep_vba=False,
            data_only=True,
            keep_links=True)
        self.ws = wb.active

    def __iter__(self):
        for row in self.ws.iter_rows(
                values_only=True, min_row=self.min_row):
            data = self._map_content(row)
            if data:
                data = self._normalize(data)
                yield data

    def _normalize(self, data):
        data = data._replace(
            title=shorten4datastore(
                normalize_whitespaces(data.title), 250),
            add_title=shorten4datastore(
                normalize_whitespaces(data.add_title), 250),
            author=shorten4datastore(
                normalize_whitespaces(data.author), 150),
            series=shorten4datastore(
                normalize_whitespaces(data.series), 250),
            publisher=shorten4datastore(
                normalize_whitespaces(data.publisher), 150),
            pub_date=shorten4datastore(
                normalize_date(data.pub_date), 25),
            pub_place=shorten4datastore(
                normalize_whitespaces(data.pub_place), 50),
            summary=shorten4datastore(
                normalize_whitespaces(data.summary), 500),
            isbn=normalize_isbn(data.isbn),
            upc=shorten4datastore(
                value2string(data.upc), 20),
            other_no=shorten4datastore(
                value2string(data.other_no), 25),
            price_list=normalize_price(data.price_list),
            price_disc=normalize_price(data.price_disc),
            desc_url=shorten4datastore(data.desc_url, 500),
            misc=shorten4datastore(
                normalize_whitespaces(data.misc), 250))
        return data

    def _map_content(self, row):
        """
        returns rows where title column has some value
        """
        try:
            kwargs = dict()
            if row[self.title_col] is not None:
                title = str(row[self.title_col]).strip()
                if title != '':
                    kwargs['title'] = title
                    if self.add_title_col is not None:
                        kwargs['add_title'] = row[self.add_title_col]
                    if self.author_col is not None:
                        kwargs['author'] = row[self.author_col]
                    if self.series_col is not None:
                        kwargs['series'] = row[self.series_col]
                    if self.publisher_col is not None:
                        kwargs['publisher'] = row[self.publisher_col]
                    if self.pub_date_col is not None:
                        kwargs['pub_date'] = row[self.pub_date_col]
                    if self.pub_place_col is not None:
                        kwargs['pub_place'] = row[self.pub_place_col]
                    if self.summary_col is not None:
                        kwargs['summary'] = row[self.summary_col]
                    if self.isbn_col is not None:
                        kwargs['isbn'] = row[self.isbn_col]
                    if self.upc_col is not None:
                        kwargs['upc'] = row[self.upc_col]
                    if self.other_no_col is not None:
                        kwargs['other_no'] = row[self.other_no_col]
                    if self.price_list_col is not None:
                        kwargs['price_list'] = row[self.price_list_col]
                    if self.price_disc_col is not None:
                        kwargs['price_disc'] = row[self.price_disc_col]
                    if self.desc_url_col is not None:
                        kwargs['desc_url'] = row[self.desc_url_col]
                    if self.misc_col is not None:
                        kwargs['misc'] = row[self.misc_col]

                    return VenData(**kwargs)
        except IndexError:
            return None


def save2spreadsheet(fh, saving_status, system, data):
    try:
        if os.path.isfile(fh):
            os.remove(fh)

        red_font = Font(color='CC0000')
        order_wb = Workbook()
        order_ws = order_wb.active
        address_line1 = f'BookOps {system}'
        address_line2 = '31-11 Thomson Avenue'
        address_line3 = 'Long Island City, NY 11101'

        # set columns width
        order_ws.column_dimensions['A'].width = 4
        order_ws.column_dimensions['C'].width = 16
        order_ws.column_dimensions['D'].width = 25
        order_ws.column_dimensions['E'].width = 20
        order_ws.column_dimensions['F'].width = 9
        order_ws.column_dimensions['G'].width = 8
        order_ws.column_dimensions['H'].width = 10
        order_ws.column_dimensions['I'].width = 12
        order_ws.column_dimensions['J'].width = 18

        order_ws.cell(row=2, column=1).value = address_line1
        order_ws.cell(row=2, column=1).font = FONT_BOLD
        order_ws.cell(row=3, column=1).value = address_line2
        order_ws.cell(row=3, column=1).font = FONT_BOLD
        order_ws.cell(row=4, column=1).value = address_line3
        order_ws.cell(row=4, column=1).font = FONT_BOLD

        # headers
        order_ws.cell(row=6, column=1).value = '#'
        order_ws.cell(row=6, column=1).fill = FILL_GRAY
        order_ws.cell(row=6, column=2).value = 'SKU'
        order_ws.cell(row=6, column=2).fill = FILL_GRAY
        order_ws.cell(row=6, column=3).value = 'ISBN'
        order_ws.cell(row=6, column=3).fill = FILL_GRAY
        order_ws.cell(row=6, column=4).value = 'Title'
        order_ws.cell(row=6, column=4).fill = FILL_GRAY
        order_ws.cell(row=6, column=5).value = 'Author'
        order_ws.cell(row=6, column=5).fill = FILL_GRAY
        order_ws.cell(row=6, column=6).value = 'Unit Price'
        order_ws.cell(row=6, column=6).fill = FILL_GRAY
        order_ws.cell(row=6, column=7).value = 'Copies'
        order_ws.cell(row=6, column=7).fill = FILL_GRAY
        order_ws.cell(row=6, column=8).value = 'Total Price'
        order_ws.cell(row=6, column=8).fill = FILL_GRAY
        order_ws.cell(row=6, column=9).value = 'o Number'
        order_ws.cell(row=6, column=9).fill = FILL_GRAY
        order_ws.cell(row=6, column=10).value = 'blanket PO'
        order_ws.cell(row=6, column=10).fill = FILL_GRAY

        r = 1
        for title in data:
            row = []
            row.append(r)
            row.extend(title)
            order_ws.append(row)
            order_ws.cell(row=6 + r, column=9).font = red_font
            order_ws.cell(row=6 + r, column=10).font = red_font
            r += 1
        last_r = 6 + r
        order_ws.cell(row=last_r + 2, column=3).value = 'total copies='
        order_ws.cell(row=last_r + 2, column=3).fill = FILL_GRAY
        order_ws.cell(row=last_r + 2, column=4).value = \
            '=SUM(G7:G' + str(last_r) + ')'
        order_ws.cell(row=last_r + 2, column=4).fill = FILL_GRAY
        order_ws.cell(row=last_r + 3, column=3).value = 'total cost='
        order_ws.cell(row=last_r + 3, column=3).fill = FILL_GRAY
        order_ws.cell(row=last_r + 3, column=4).value = \
            '=SUM(H7:H' + str(last_r) + ')'
        order_ws.cell(row=last_r + 3, column=4).fill = FILL_GRAY

        order_wb.save(filename=fh)
        saving_status.set('Data saved to spreadsheet successfuly.')

    except WindowsError as e:
        raise BabelError(e)

    except Exception as exc:
        _, _, exc_traceback = sys.exc_info()
        tb = format_traceback(exc, exc_traceback)
        mlogger.error(
            'Unhandled error on saving to sheet.'
            f'Traceback: {tb}')
        raise BabelError(exc)
