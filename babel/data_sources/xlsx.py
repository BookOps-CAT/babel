# xlsx spreadsheet parser
# how about list of namedtuples as output parsing spreadsheet? pros: immutable, access by name, default values

from collections import namedtuple

from openpyxl import load_workbook
from openpyxl import Workbook


Data = namedtuple(
    'Data',
    [
        'title',
        'author',
        'series',
        'publisher',
        'pub_date',
        'summary',
        'isbn',
        'upc',
        'other_no',
        'price_list',
        'price_disc',
        'desc_url'
    ],
    defaults=[
        None, None, None, None,
        None, None, None, None,
        None, 0.0, 0.0, None])


class OrderDataReader:
    """
    Parses specified rows and columns of the xlsx spreadsheet
    The reader can be iterater over in a loop

    Arguments:
    ----------
    fh: str
        path to xlsx spreadsheet
    header_row: int
        row number of a header, index starts with 0
    title_col: int
        title column number, index starts with 0
    author_col: int
        author column number, index starts with 0
    series_col: int
        series column number, index starts with 0
    publisher_col: int
        publisher column number, index starts with 0
    pub_datae_col: int
        publicattion column number, index starts with 0
    summary_col: int
        summary column number, index starts with 0
    isbn_col: int
        ISBN column number, index starts with 0
    upc_col: int
        UPC column number, index starts with 0
    other_no_col: int
        publisher interal ID, etc. column number, index starts with 0
    price_list_col: int
        list price column number, index starts with 0
    price_disc: int
        discounted price column number, index starts with 0
    desc_url: int
        URL of description column number, index starts with 0

    """

    def __init__(
            self, fh,
            header_row=None,
            title_col=None,
            author_col=None,
            series_col=None,
            publisher_col=None,
            pub_date_col=None,
            summary_col=None,
            isbn_col=None,
            upc_col=None,
            other_no_col=None,
            price_list_col=None,
            price_disc_col=None,
            desc_url_col=None):

        self.min_row = header_row
        self.header_row = header_row
        self.title_col = title_col
        self.author_col = author_col
        self.series_col = series_col
        self.publisher_col = publisher_col
        self.pub_date_col = pub_date_col
        self.summary_col = summary_col
        self.isbn_col = isbn_col
        self.upc_col = upc_col
        self.other_no_col = other_no_col
        self.price_list_col = price_list_col
        self.price_disc_col = price_disc_col
        self.url_col = url_col

        if not header_row:
            raise AttributeError(
                'Header row number is a required argument')

        if not title_col:
            raise AttributeError(
                'Title column letter is a required argument')

        wb = load_workbook(
            filename=fh,
            read_only=True,
            keep_vba=False,
            data_only=True,
            keep_links=True)
        self.ws = wb.active

    def __iter__(self):
        for row in self.ws.iter_rows(values_only=True, min_row=self.min_row):
            data = self._map2content(row)
            if data:
                yield data

    def _map2content(self, row):
        """
        returns rows where title column has some value
        """
        kwargs = dict()
        if row[self.title_col]:
            kwargs['title'] = row[self.title_col]
            if self.author_col:
                kwargs['author'] = row[self.author_col]
            if self.series_col:
                kwargs['series'] = row[self.series_col]
            if self.publisher_col:
                kwargs['publisher'] = row[self.publisher_col]
            if self.pub_date_col:
                kwargs['pub_date'] = row[self.pub_date_col]
            if self.summary_col:
                kwargs['summary'] = row[self.summary_col]
            if self.isbn_col:
                kwargs['isbn'] = row[self.isbn_col]
            if self.upc_col:
                kwargs['upc'] = row[self.upc_col]
            if self.other_no_col:
                kwargs['other_no'] = row[self.other_no_col]
            if self.price_list_col:
                kwargs['price_list'] = float(row[self.price_list_col])
            if self.price_disc_col:
                kwargs['price_disc'] = float(row[self.price_disc_col])
            if self.desc_url_col:
                kwargs['desc_url'] = row[self.desc_url]

            return Data(**kwargs)
