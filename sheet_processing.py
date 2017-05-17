# xlsx reader & writer

from openpyxl import load_workbook, Workbook
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.styles import Font, PatternFill
import os.path
import logging
import datetime

from convert_price import dollars2cents, cents2dollars
from celldata_parser import parse_isbn, parse_year
from Z3950_communicator import query


module_logger = logging.getLogger('babel_logger.sheet')


FONT_BOLD = Font(bold=True)
FILL_GRAY = PatternFill(fill_type='solid',
                        start_color='C4C5C6',
                        end_color='C4C5C6')




class SheetManipulator:

    def __init__(self, file):
        # define styles for writing sheet
        self.font = Font(bold=True)
        self.red_font = Font(bold=True, color='CC0000')
        self.white_font = Font(color='FFFFFF')

        self.wb = load_workbook(filename=file,
                                data_only=True,
                                guess_types=False)
        self.ws = self.wb.active
        self.max_r = 0
        self.max_c = 1
        for row in self.ws.rows:
            self.max_r += 1
            max_c = len(row)
            if max_c > self.max_c:
                self.max_c = max_c
        self.range = 'A1:' + str(
            get_column_letter(self.max_c)) + str(self.max_r)
        self.last_col = get_column_letter(self.max_c)

    def get_column_letters(self):
        column_letters = []
        for column in range(1, self.max_c + 1):
            column_letters.append(get_column_letter(column))
        return column_letters

    def extract_data(self):
        self.sheet_data = []
        # find last column & last row (column not accurate in openpyxl)
        # for row in tuple(self.ws.iter_rows(self.range)):
        for row in self.ws[self.range]:
            row_data = []
            for cell in row:
                row_data.append(cell.value)
            self.sheet_data.append(row_data)
        return self.sheet_data

    def cart_sheet(self, fname, **kwargs):

        self.head_row = self.ws[
            'A' + str(kwargs['head_row']) +
            ':' + self.last_col + str(kwargs['head_row'])]
        self.data_range = 'A' + str(
            kwargs['head_row'] + 1) + ':' + str(
            get_column_letter(self.max_c)) + str(
            self.max_r + 1)

        # define cart sheet
        self.new_wb = Workbook()
        self.new_ws_cart = self.new_wb.active
        self.new_ws_cart.title = 'cart'

        # define meta sheet
        self.new_ws_meta = self.new_wb.create_sheet('meta')
        self.new_ws_meta.protection.set_password('babel')

        # reference legend area
        dr = 1
        # print kwargs['distrDetails']
        distrLegend = kwargs['distrDetails'].split('\n')
        for line in distrLegend[1:]:
            self.new_ws_cart.cell(row=dr, column=1).value = line
            self.new_ws_cart.cell(row=dr, column=1).font = self.font
            dr += 1
        audnLegend = [
            'empty cell = adult',
            'j = juvenile',
            'y = young adult']
        dr += 1
        self.new_ws_cart.cell(row=dr, column=1).value = '; '.join(
            audnLegend)
        self.new_ws_cart.cell(row=dr, column=1).font = self.font

        c = 1
        for key, value in kwargs['codeTotalQntBranch'].items():
            self.new_ws_meta.cell(row=2, column=c).value = key
            self.new_ws_meta.cell(row=3, column=c).value = value[0]
            self.new_ws_meta.cell(row=4, column=c).value = value[1]
            c += 1
        # codes_range must be in A1 notation
        codes_range = 'meta!A2:%s4' % (get_column_letter(c - 1))
        # headings row
        extra_columns = ()

        if kwargs['z3950target'] is None:
            target_name = 'no target'
        else:
            target_name = kwargs['z3950target']['name']

        standard_columns = (
            'Distribution',
            'Audience',
            'PO per line',
            target_name,
            'Branches',  # optional ask if needed
            'Total Qty',
            'Total Price')  # optional ask if needed
        extra_columns = extra_columns + kwargs['collabs']
        if kwargs['priceDefault'] != 0.0 or kwargs['priceDisc_col'] == '':
            extra_columns = extra_columns + ('Unit Price', )
            shifted_price_col = get_column_letter(len(extra_columns))
        else:
            price_col = kwargs['priceDisc_col']
            shifted_price_col_num = column_index_from_string(price_col) \
                + len(extra_columns) + len(standard_columns)
            shifted_price_col = get_column_letter(shifted_price_col_num)
        distr_col = len(extra_columns) + 1
        # branches_col = len(extra_columns) + 5
        total_qty_col = len(extra_columns) + 6  # leaves space for fixed cols
        total_price_col = len(extra_columns) + 7
        extra_columns = extra_columns + standard_columns
        shift_index = len(extra_columns)

        # add sheet totals
        head_row = dr + 6
        # totals area
        self.new_ws_cart.cell(row=dr + 2, column=1).value = '# of titles='
        self.new_ws_cart.cell(row=dr + 2, column=1).font = self.font
        self.new_ws_cart.cell(row=dr + 3, column=1).value = '# of copies='
        self.new_ws_cart.cell(row=dr + 3, column=1).font = self.font
        self.new_ws_cart.cell(row=dr + 4, column=1).value = 'total price='
        self.new_ws_cart.cell(row=dr + 4, column=1).font = self.font
        self.new_ws_cart.cell(row=dr + 2, column=2).value = \
            "=COUNTA(%s%s:%s1000)" % (
            get_column_letter(distr_col),
            head_row + 1,
            get_column_letter(distr_col))
        self.new_ws_cart.cell(row=dr + 2, column=2).font = self.red_font
        self.new_ws_cart.cell(row=dr + 3, column=2).value = \
            '=SUMIF(%s%s:%s1000, ">0")' % (
            get_column_letter(total_qty_col),
            head_row + 1,
            get_column_letter(total_qty_col))
        self.new_ws_cart.cell(row=dr + 3, column=2).font = self.red_font
        self.new_ws_cart.cell(row=dr + 4, column=2).value = \
            '=SUMIF(%s%s:%s1000, ">0")' % (
            get_column_letter(total_price_col),
            head_row + 1,
            get_column_letter(total_price_col))
        self.new_ws_cart.cell(row=dr + 4, column=2).font = self.red_font
        self.new_ws_cart.cell(row=dr + 5, column=1).value = None

        new_headings = []
        for col in extra_columns:
            new_headings.append(col)
        for heading in self.head_row:
            for data in heading:
                new_headings.append(data.value)
        self.new_ws_cart.append(new_headings)

        # remaining rows
        # for row in tuple(self.ws.iter_rows(self.data_range)):
        row_counter = 0
        for row in self.ws[self.data_range]:
            row_counter += 1
            # branch formula recognizes only 2 codes in distribution
            totQnt_formula = '=HlOOKUP(LEFT(TRIM(INDIRECT("R"&ROW()&"C%s", FALSE)),1),%s,2,FALSE)+IF(LEN(INDIRECT("R"&ROW()&"C%s",FALSE))=2,HLOOKUP(RIGHT(TRIM(INDIRECT("R"&ROW()&"C%s",FALSE)),1),%s,2,FALSE),0)' % (
                distr_col, codes_range, distr_col, distr_col, codes_range)
            totPrice_formula = '=HLOOKUP(LEFT(TRIM(INDIRECT("R"&ROW()&"C%s", FALSE)),1),%s,2,FALSE)*INDIRECT("%s"&ROW())+IF(LEN(INDIRECT("R"&ROW()&"C%s",FALSE))=2,HLOOKUP(RIGHT(TRIM(INDIRECT("R"&ROW()&"C%s",FALSE)),1),%s,2,FALSE)*INDIRECT("%s"&ROW()),0)' % (
                distr_col, codes_range, shifted_price_col,
                distr_col, distr_col, codes_range, shifted_price_col)
            branch_formula = '=CONCATENATE((HlOOKUP(LEFT(TRIM(INDIRECT("R"&ROW()&"C%s", FALSE)),1),%s,3,FALSE)),(IF(LEN(INDIRECT("R"&ROW()&"C%s",FALSE))=2,HLOOKUP(RIGHT(TRIM(INDIRECT("R"&ROW()&"C%s",FALSE)),1),%s,3,FALSE),"")))' % (
                distr_col, codes_range, distr_col, distr_col, codes_range)
            new_values = []
            col_counter = 0
            for col in extra_columns:
                col_counter += 1
                if col == 'Total Qty':
                    new_values.append(totQnt_formula)
                elif col == 'Total Price':
                    new_values.append(totPrice_formula)
                elif col == 'Branches':
                    new_values.append(branch_formula)
                elif col == 'Unit Price':
                    if kwargs['priceDefault'] != 0.0:
                        priceDefault = dollars2cents(kwargs['priceDefault'])
                        priceDefault = cents2dollars(priceDefault)
                        new_values.append(priceDefault)
                    else:
                        discount = kwargs['discount']
                        discount = str(discount).zfill(2)
                        discount = '0.' + discount
                        discount = float(discount)
                        # kwargs['priceReg_col']
                        list_price = self.ws[
                            kwargs['priceReg_col'] +
                            str(kwargs['head_row'] + row_counter)
                        ].value

                        if list_price is None or type(list_price) is not float:
                            list_price = 0.0

                        disc_price = list_price - (list_price * discount)
                        disc_price = dollars2cents(disc_price)
                        disc_price = cents2dollars(disc_price)
                        new_values.append(disc_price)
                elif col == 'Audience':
                    audn_col = col_counter
                    new_values.append(None)
                elif col == 'PO per line':
                    po_per_line_col = col_counter
                    new_values.append(None)
                elif col == target_name:
                    if kwargs['isbn_col'] != '':
                        isbn_value = self.ws[
                            kwargs['isbn_col'] +
                            str(kwargs['head_row'] + row_counter)].value
                        isbn = parse_isbn(isbn_value)
                        if isbn is not None and \
                                kwargs['z3950target'] is not None:
                            catalog_query = query(
                                target=kwargs['z3950target'],
                                keyword=isbn,
                                qualifier='isbn')
                            if catalog_query[0]:
                                if len(catalog_query[1]) > 0:
                                    query_result = 'found'
                                else:
                                    query_result = None
                        else:
                            query_result = None
                    else:
                        query_result = None
                    new_values.append(query_result)
                else:
                    new_values.append(None)
            old_values = []
            for cell in row:
                try:
                    content = cell.value.strip().replace(
                        '\n', ' ').replace('\t', ' ')
                except:
                    content = cell.value
                old_values.append(content)
            new_values.extend(old_values)
            self.new_ws_cart.append(new_values)

        # record new column & rows structure
        pre_shift_metadata = kwargs.copy()
        shifted_metadata = {}
        columns = ('title_col', 'author_col',
                   'isbn_col', 'venNum_col',
                   'publisher_col', 'pubDate_col',
                   'pubPlace_col', 'priceReg_col')
        for key, value in pre_shift_metadata.items():
            if key in columns:
                if value != '':
                    shifted_metadata[key] = column_index_from_string(
                        value) + shift_index
                elif value == '':
                    shifted_metadata[key] = value

        shifted_metadata['vendorSheetTemplate_id'] = kwargs['sheetTemp_id']
        shifted_metadata['distrTemplate_id'] = kwargs['distr_id']
        shifted_metadata['head_row'] = head_row
        shifted_metadata['distr_col'] = distr_col
        shifted_metadata['audn_col'] = audn_col
        shifted_metadata['po_per_line_col'] = po_per_line_col
        shifted_metadata['priceDisc_col'] = column_index_from_string(
            shifted_price_col)

        # add sheet metadata to meta sheet
        meta_str = ''
        for key, value in shifted_metadata.items():
            meta_str = meta_str + key + '=' + str(value) + ';'
        self.new_ws_meta.cell(row=1, column=1).value = meta_str

        # determine file name
        n = 0
        fh = fname + '.xlsx'
        while os.path.isfile(fh):
            n += 1
            fh = fname + '(' + str(n) + ')' + '.xlsx'
        else:
            filehandle = fh
        self.new_wb.save(filename=filehandle)
        return filehandle

    def extract_meta(self):
        self.ws_meta = self.wb['meta']
        meta_str = self.ws_meta['A1'].value
        meta_lst = meta_str.split(';')
        self.metadata = {}
        for item in meta_lst:
            pos = item.find('=')
            key = item[:pos]
            value = item[pos + 1:]
            if value != '':
                value = int(value)
                if key in (
                        'title_col', 'distr_col', 'audn_col',
                        'po_per_line_col', 'author_col',
                        'isbn_col', 'venNum_col', 'publisher_col',
                        'pubDate_col', 'pubPlace_col', 'priceReg_col',
                        'priceDisc_col'):
                    value = get_column_letter(value)
            self.metadata[key] = value
        self.start_row = self.metadata['head_row']
        return self.metadata

    def find_orders(self, distr_codes):
        self.orders = []
        self.total_cost = 0.0
        self.distr_codes = distr_codes
        self.ws_orders = self.wb['cart']
        self.order_range = 'A' + str(
            self.start_row + 1) + ':' + str(
            get_column_letter(self.max_c)) + str(
            self.max_r + 1)
        distr_col_num = column_index_from_string(self.metadata['distr_col'])

        selected_rows = []
        r = self.metadata['head_row'] + 1

        for row in self.ws_orders[self.order_range]:
            cell_num = 1
            for cell in row:
                if cell_num == distr_col_num:
                    codes_qty = 0
                    if cell.value is not None:
                        for code in str(cell.value):
                            try:
                                code = code.upper()
                            except:
                                pass
                            if code in self.distr_codes:
                                codes_qty += 1
                        # append only if all codes in cel match self.distr_codes
                        if codes_qty == len(str(cell.value)):
                            selected_rows.append(r)
                            break
                cell_num += 1
            r += 1

        for row_num in selected_rows:
            order = {}
            for key, value in self.metadata.items():
                if key in (
                    'title_col', 'distr_col', 'audn_col',
                    'po_per_line_col', 'author_col',
                    'isbn_col', 'venNum_col', 'publisher_col',
                    'pubDate_col', 'pubPlace_col', 'priceReg_col',
                    'priceDisc_col'
                ):
                    if value != '':
                        cell = self.ws_orders[
                            self.metadata[key] + str(row_num)]
                        content = cell.value
                        order[key] = content

                        if key == 'priceDisc_col':
                            self.total_cost = self.total_cost + content
                        if key == 'distr_col':
                            try:
                                order['distr_col'] = order['distr_col'].upper()
                            except:
                                order['distr_col'] = order['distr_col']                
                    else:
                        order[key] = None

            self.orders.append(order)
            self.order_qty = len(selected_rows)
        # summary can serve to check if selected = added to db
        # at this moment did not pursue it
        return {'summary': None, 'orders': self.orders}


def create_order(fname, library, data):
    red_font = Font(color='CC0000')
    order_wb = Workbook()
    order_ws = order_wb.active
    # title_count = 0
    address_line1 = 'BookOps %s' % library
    address_line2 = '31-11 Thomson Avenue'
    address_line3 = 'Long Island City, NY 11101'

    # set columns width
    order_ws.column_dimensions['A'].width = 4
    order_ws.column_dimensions['C'].width = 16
    order_ws.column_dimensions['D'].width = 25
    order_ws.column_dimensions['E'].width = 20
    order_ws.column_dimensions['F'].width = 9
    order_ws.column_dimensions['G'].width = 8
    order_ws.column_dimensions['H'].width = 10
    order_ws.column_dimensions['I'].width = 12
    order_ws.column_dimensions['J'].width = 18

    order_ws.cell(row=2, column=1).value = address_line1
    order_ws.cell(row=2, column=1).font = FONT_BOLD
    order_ws.cell(row=3, column=1).value = address_line2
    order_ws.cell(row=3, column=1).font = FONT_BOLD
    order_ws.cell(row=4, column=1).value = address_line3
    order_ws.cell(row=4, column=1).font = FONT_BOLD

    # headers
    order_ws.cell(row=6, column=1).value = '#'
    order_ws.cell(row=6, column=1).fill = FILL_GRAY
    order_ws.cell(row=6, column=2).value = 'SKU'
    order_ws.cell(row=6, column=2).fill = FILL_GRAY
    order_ws.cell(row=6, column=3).value = 'ISBN'
    order_ws.cell(row=6, column=3).fill = FILL_GRAY
    order_ws.cell(row=6, column=4).value = 'Title'
    order_ws.cell(row=6, column=4).fill = FILL_GRAY
    order_ws.cell(row=6, column=5).value = 'Author'
    order_ws.cell(row=6, column=5).fill = FILL_GRAY
    order_ws.cell(row=6, column=6).value = 'Unit Price'
    order_ws.cell(row=6, column=6).fill = FILL_GRAY
    order_ws.cell(row=6, column=7).value = 'Copies'
    order_ws.cell(row=6, column=7).fill = FILL_GRAY
    order_ws.cell(row=6, column=8).value = 'Total Price'
    order_ws.cell(row=6, column=8).fill = FILL_GRAY
    order_ws.cell(row=6, column=9).value = 'o Number'
    order_ws.cell(row=6, column=9).fill = FILL_GRAY
    order_ws.cell(row=6, column=10).value = 'blanket PO'
    order_ws.cell(row=6, column=10).fill = FILL_GRAY

    r = 1
    for title in data:
        row = []
        row.append(r)
        row.extend(title)
        order_ws.append(row)
        order_ws.cell(row=6 + r, column=9).font = red_font
        order_ws.cell(row=6 + r, column=10).font = red_font
        r += 1
    last_r = 6 + r
    order_ws.cell(row=last_r + 2, column=3).value = 'total copies='
    order_ws.cell(row=last_r + 2, column=3).fill = FILL_GRAY
    order_ws.cell(row=last_r + 2, column=4).value = \
        '=SUM(G7:G' + str(last_r) + ')'
    order_ws.cell(row=last_r + 2, column=4).fill = FILL_GRAY
    order_ws.cell(row=last_r + 3, column=3).value = 'total cost='
    order_ws.cell(row=last_r + 3, column=3).fill = FILL_GRAY
    order_ws.cell(row=last_r + 3, column=4).value = \
        '=SUM(H7:H' + str(last_r) + ')'
    order_ws.cell(row=last_r + 3, column=4).fill = FILL_GRAY

    # determine if new file name needed
    n = 0
    fh = fname + '.xlsx'
    while os.path.isfile(fh):
            n += 1
            fh = fname + '(' + str(n) + ')' + '.xlsx'
    else:
        filehandle = fh
    try:
        order_wb.save(filename=filehandle)
        return filehandle
    except:
        return None


def export_search(fh, data):
    # saves search results to Excel spreadsheet
    result_wb = Workbook()
    result_ws = result_wb.active

    headings = [
        'library',
        'language',
        'vendor',
        'mat type',
        'title',
        'title trans.',
        'author',
        'author trans.',
        'ISBN',
        'vendor #',
        'pub place',
        'publisher',
        'pub date',
        'qty',
        'disc. price',
        'blanket PO',
        'wlo #',
        'oNumber',
        'bNumber',
        'locations',
        'selector',
        'order date'
    ]

    # headings row
    c = 1
    for heading in headings:
        result_ws.cell(row=1, column=c).value = heading
        result_ws.cell(row=1, column=c).font = FONT_BOLD
        result_ws.cell(row=1, column=c).fill = FILL_GRAY
        c += 1

    col_short = ['A', 'B', 'D', 'F', 'H', 'J', 'K', 'M', 'N', 'O']
    for col in col_short:
        result_ws.column_dimensions[col].width = 8
    col_long = ['E', 'G', 'I', 'T', 'V']
    for col in col_long:
        result_ws.column_dimensions[col].width = 15
    col_medium = ['C', 'P', 'Q', 'R', 'S', 'U']
    for col in col_medium:
        result_ws.column_dimensions[col].width = 12

    r = 2
    for order in data:
        for key, value in order.iteritems():
            if key == 'library':
                c = 1
            if key == 'lang':
                c = 2
            if key == 'vendor':
                c = 3
            if key == 'matType':
                c = 4
            if key == 'title':
                c = 5
            if key == 'title_trans':
                c = 6
            if key == 'author':
                c = 7
            if key == 'author_trans':
                c = 8
            if key == 'isbn':
                c = 9
            if key == 'venNo':
                c = 10
            if key == 'pubPlace':
                c = 11
            if key == 'publisher':
                c = 12
            if key == 'pubDate':
                c = 13
            if key == 'qty':
                c = 14
            if key == 'priceDisc':
                c = 15
            if key == 'blanketPO':
                c = 16
            if key == 'wlo_id':
                c = 17
            if key == 'oNumber':
                c = 18
            if key == 'bNumber':
                c = 19
            if key == 'locations':
                c = 20
                value = ','.join(value)
            if key == 'selector':
                c = 21
            if key == 'date':
                c = 22
                value = value.strftime('%d/%m/%Y')

            result_ws.cell(row=r, column=c).value = value
        r += 1

    try:
        result_wb.save(filename=fh)
        return fh
    except IOError:
        module_logger.exception('Search export error')
        return None
